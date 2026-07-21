"""Servicios para importación masiva de movimientos de inventario.

Soporta archivos CSV/XLS/XLSX, normaliza encabezados, construye una vista
previa de validación y confirma el lote en una sola transacción.
"""

from __future__ import annotations

import csv
from datetime import date, datetime
from difflib import get_close_matches
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
import re
import unicodedata

from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
import xlrd

from apps.accounts.models import AppUser
from apps.inventory.models import MateriaPrima, Proveedor, TrabajadorProduccion
from apps.inventory.serializers import MovimientoInventarioSerializer


TEMPLATE_HEADERS = [
    "materia_prima",
    "tipo",
    "cantidad",
    "fecha",
    "proveedor",
    "trabajador",
    "motivo",
    "referencia",
]

CANONICAL_HEADERS = set(TEMPLATE_HEADERS)

# Acepta variaciones reales de encabezados para reducir rechazos por formato.
HEADER_ALIASES = {
    "materia_prima": "materia_prima",
    "materiaprima": "materia_prima",
    "materia prima": "materia_prima",
    "materia-prima": "materia_prima",
    "materia prima nombre": "materia_prima",
    "nombre materia prima": "materia_prima",
    "material": "materia_prima",
    "tipo": "tipo",
    "tipo movimiento": "tipo",
    "cantidad": "cantidad",
    "cantidad movimiento": "cantidad",
    "fecha": "fecha",
    "fecha movimiento": "fecha",
    "proveedor": "proveedor",
    "proveedor_nombre": "proveedor",
    "nombre proveedor": "proveedor",
    "trabajador": "trabajador",
    "trabajador produccion": "trabajador",
    "trabajador_produccion": "trabajador",
    "nombre trabajador": "trabajador",
    "trabajador_codigo": "trabajador",
    "codigo_trabajador": "trabajador",
    "responsable": "trabajador",
    "motivo": "motivo",
    "referencia": "referencia",
    "codigo referencia": "referencia",
}

DATE_FORMATS = (
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%Y/%m/%d",
)


def _normalize_text(value: object) -> str:
    """Normaliza texto para comparaciones tolerantes a mayúsculas/acentos."""

    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text).casefold()


def _normalize_lookup_key(value: object) -> str:
    """Genera una clave de búsqueda estable removiendo ruido común."""

    normalized = _normalize_text(value)
    normalized = re.sub(r"[-_/\\.]+", " ", normalized)
    normalized = re.sub(r"[^a-z0-9\s]", "", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def _compact_lookup_key(value: object) -> str:
    """Versión compacta de la clave para matching flexible."""

    return _normalize_lookup_key(value).replace(" ", "")


def _build_header_observation(raw_text: str, canonical: str) -> str | None:
    """Informa cuándo una cabecera fue reinterpretada por alias/fuzzy."""

    if not raw_text:
        return None
    if _normalize_lookup_key(raw_text) == _normalize_lookup_key(canonical):
        return None
    return f"Se interpretó '{raw_text}' como '{canonical}'."


def _resolve_header(value: object) -> tuple[str, str | None]:
    """Resuelve cabecera de archivo hacia la cabecera canónica del sistema."""

    raw_text = _cell_to_text(value)
    # Estrategia en capas: exacto -> alias -> fuzzy para soportar plantillas ruidosas.
    candidates = {
        _normalize_text(raw_text),
        _normalize_lookup_key(raw_text),
        _compact_lookup_key(raw_text),
    }

    for candidate in list(candidates):
        if not candidate:
            continue

        variants = {
            candidate,
            candidate.replace("-", "_"),
            candidate.replace(" ", "_"),
            candidate.replace("_", " "),
            candidate.replace("_", ""),
        }

        for variant in variants:
            if variant in CANONICAL_HEADERS:
                return variant, _build_header_observation(raw_text, variant)
            if variant in HEADER_ALIASES:
                canonical = HEADER_ALIASES[variant]
                return canonical, _build_header_observation(raw_text, canonical)

    alias_candidates = set(HEADER_ALIASES) | CANONICAL_HEADERS
    normalized_target = _normalize_lookup_key(raw_text)
    compact_target = _compact_lookup_key(raw_text)

    normalized_matches = get_close_matches(normalized_target, alias_candidates, n=1, cutoff=0.72)
    if normalized_matches:
        match = normalized_matches[0]
        canonical = HEADER_ALIASES.get(match, match)
        return canonical, _build_header_observation(raw_text, canonical)

    compact_aliases = {
        _compact_lookup_key(alias): HEADER_ALIASES.get(alias, alias)
        for alias in alias_candidates
    }
    compact_matches = get_close_matches(compact_target, list(compact_aliases), n=1, cutoff=0.8)
    if compact_matches:
        canonical = compact_aliases[compact_matches[0]]
        return canonical, _build_header_observation(raw_text, canonical)

    fallback = _normalize_text(raw_text).replace("-", "_").replace(" ", "_")
    canonical = HEADER_ALIASES.get(fallback, fallback)
    return canonical, _build_header_observation(raw_text, canonical)


def _cell_to_text(value: object) -> str:
    """Convierte cualquier celda a representación textual consistente."""

    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _parse_decimal(value: object) -> Decimal:
    """Parsea cantidades permitiendo distintos formatos numéricos."""

    if value is None or str(value).strip() == "":
        raise ValueError("La cantidad es obligatoria.")

    if isinstance(value, Decimal):
        return value

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip().replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")

    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError("La cantidad debe ser numérica.") from exc


def _parse_date(value: object) -> str:
    """Parsea fecha y retorna formato ISO para persistencia."""

    if value is None or str(value).strip() == "":
        raise ValueError("La fecha es obligatoria.")

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    for date_format in DATE_FORMATS:
        try:
            return datetime.strptime(text, date_format).date().isoformat()
        except ValueError:
            continue

    raise ValueError("La fecha debe usar formato YYYY-MM-DD o DD/MM/YYYY.")


def _read_csv_rows(file_bytes: bytes) -> tuple[list[tuple[int, dict[str, object]]], list[str]]:
    """Lee un CSV y lo transforma al formato interno fila/campos."""

    decoded_text: str
    try:
        decoded_text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        decoded_text = file_bytes.decode("latin-1")

    reader = csv.reader(StringIO(decoded_text))
    rows = list(reader)
    return _rows_from_matrix(rows)


def _read_openpyxl_rows(file_bytes: bytes) -> tuple[list[tuple[int, dict[str, object]]], list[str]]:
    """Lee un archivo XLSX/XLSM con openpyxl."""

    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    return _rows_from_matrix(rows)


def _read_xls_rows(file_bytes: bytes) -> tuple[list[tuple[int, dict[str, object]]], list[str]]:
    """Lee un archivo XLS legado usando xlrd."""

    workbook = xlrd.open_workbook(file_contents=file_bytes)
    worksheet = workbook.sheet_by_index(0)
    rows: list[list[object]] = []

    for row_index in range(worksheet.nrows):
        row_values: list[object] = []
        for column_index in range(worksheet.ncols):
            cell = worksheet.cell(row_index, column_index)
            value: object = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                value = xlrd.xldate.xldate_as_datetime(cell.value, workbook.datemode)
            row_values.append(value)
        rows.append(row_values)

    return _rows_from_matrix(rows)


def _rows_from_matrix(rows: list[list[object]]) -> tuple[list[tuple[int, dict[str, object]]], list[str]]:
    """Mapea matriz cruda a filas numeradas con cabeceras normalizadas."""

    # Ignora filas vacías para estabilizar la numeración de errores por archivo.
    non_empty_rows = [row for row in rows if any(_cell_to_text(cell) for cell in row)]
    if not non_empty_rows:
        return [], []

    header_row = non_empty_rows[0]
    resolved_headers = [_resolve_header(cell) for cell in header_row]
    headers = [canonical for canonical, _observation in resolved_headers]
    header_observations = list(dict.fromkeys(observation for _canonical, observation in resolved_headers if observation))
    parsed_rows: list[tuple[int, dict[str, object]]] = []

    for row_offset, row in enumerate(non_empty_rows[1:], start=2):
        # Se rellena con None cuando faltan columnas para mantener mapeo por cabecera.
        values = row + [None] * max(0, len(headers) - len(row))
        parsed_rows.append(
            (
                row_offset,
                {
                    header: values[index] if index < len(values) else None
                    for index, header in enumerate(headers)
                    if header
                },
            )
        )

    return parsed_rows, header_observations


def _read_source_rows(file_name: str, file_bytes: bytes) -> tuple[list[tuple[int, dict[str, object]]], list[str]]:
    """Despacha el lector según extensión de archivo cargado."""

    extension = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""

    if extension == "csv":
        return _read_csv_rows(file_bytes)
    if extension in {"xlsx", "xlsm"}:
        return _read_openpyxl_rows(file_bytes)
    if extension == "xls":
        return _read_xls_rows(file_bytes)

    raise ValueError("El archivo debe estar en formato .xlsx, .xls o .csv.")


def _register_lookup_value(lookup: dict[str, object], aliases: dict[str, str], raw_value: object, instance: object) -> None:
    """Registra múltiples variantes de un valor para resolución flexible."""

    display_value = _cell_to_text(raw_value)
    keys = {
        _normalize_text(raw_value),
        _normalize_lookup_key(raw_value),
        _compact_lookup_key(raw_value),
    }

    for key in keys:
        if not key:
            continue
        lookup.setdefault(key, instance)
        aliases.setdefault(key, display_value)


def _find_lookup_value(raw_value: object, lookup: dict[str, object]) -> object | None:
    """Busca una referencia de catálogo usando distintas normalizaciones."""

    for key in (_normalize_text(raw_value), _normalize_lookup_key(raw_value), _compact_lookup_key(raw_value)):
        if key and key in lookup:
            return lookup[key]
    return None


def _find_closest_alias(raw_value: object, aliases: dict[str, str]) -> str | None:
    """Sugiere el alias más cercano para mejorar mensajes de validación."""

    compact_target = _compact_lookup_key(raw_value)
    if not compact_target:
        return None

    compact_candidates = [key for key in aliases if key and " " not in key]
    matches = get_close_matches(compact_target, compact_candidates, n=1, cutoff=0.8)
    if matches:
        return aliases[matches[0]]

    normalized_target = _normalize_lookup_key(raw_value)
    normalized_candidates = [key for key in aliases if key and " " in key]
    matches = get_close_matches(normalized_target, normalized_candidates, n=1, cutoff=0.8)
    if matches:
        return aliases[matches[0]]

    return None


def _build_catalog_error(entity_label: str, suggestion: str | None, generic_message: str) -> str:
    """Construye mensajes de error enriquecidos con sugerencia opcional."""

    if suggestion:
        return f"{generic_message} ¿Quisiste decir '{suggestion}'?"
    return generic_message


def _build_lookup_maps() -> tuple[
    dict[str, MateriaPrima],
    dict[str, str],
    dict[str, Proveedor],
    dict[str, str],
    dict[str, TrabajadorProduccion],
    dict[str, str],
]:
    """Precarga catálogos de referencia para validar filas eficientemente."""

    # Precarga catálogos en memoria para validar cientos de filas en O(1) por campo.
    materias: dict[str, MateriaPrima] = {}
    materia_aliases: dict[str, str] = {}
    proveedores: dict[str, Proveedor] = {}
    proveedor_aliases: dict[str, str] = {}
    trabajadores: dict[str, TrabajadorProduccion] = {}
    trabajador_aliases: dict[str, str] = {}

    for materia in MateriaPrima.objects.select_related("unidad_medida").all():
        if materia.nombre:
            _register_lookup_value(materias, materia_aliases, materia.nombre, materia)

    for proveedor in Proveedor.objects.all():
        if proveedor.nombre:
            _register_lookup_value(proveedores, proveedor_aliases, proveedor.nombre, proveedor)

    for trabajador in TrabajadorProduccion.objects.all():
        if trabajador.codigo_trabajador:
            _register_lookup_value(trabajadores, trabajador_aliases, trabajador.codigo_trabajador, trabajador)
        if trabajador.nombre:
            _register_lookup_value(trabajadores, trabajador_aliases, trabajador.nombre, trabajador)

    return materias, materia_aliases, proveedores, proveedor_aliases, trabajadores, trabajador_aliases


def _flatten_serializer_errors(detail: object) -> list[str]:
    """Aplana errores anidados de serializer a una lista simple de mensajes."""

    if isinstance(detail, dict):
        messages: list[str] = []
        for value in detail.values():
            messages.extend(_flatten_serializer_errors(value))
        return messages
    if isinstance(detail, list):
        messages: list[str] = []
        for item in detail:
            messages.extend(_flatten_serializer_errors(item))
        return messages
    return [str(detail)]


def _build_row_preview(
    row_number: int,
    raw_row: dict[str, object],
    materias: dict[str, MateriaPrima],
    materia_aliases: dict[str, str],
    proveedores: dict[str, Proveedor],
    proveedor_aliases: dict[str, str],
    trabajadores: dict[str, TrabajadorProduccion],
    trabajador_aliases: dict[str, str],
) -> tuple[dict[str, object], dict[str, object] | None]:
    """Valida una fila de importación y retorna preview + payload persistible."""

    # La vista previa conserva valores originales del archivo para explicar errores al usuario.
    display_values = {
        "materia_prima": _cell_to_text(raw_row.get("materia_prima")),
        "tipo": _cell_to_text(raw_row.get("tipo")),
        "cantidad": _cell_to_text(raw_row.get("cantidad")),
        "fecha": _cell_to_text(raw_row.get("fecha")),
        "proveedor": _cell_to_text(raw_row.get("proveedor")),
        "trabajador": _cell_to_text(raw_row.get("trabajador")),
        "motivo": _cell_to_text(raw_row.get("motivo")),
        "referencia": _cell_to_text(raw_row.get("referencia")),
    }
    errors: list[str] = []

    materia_nombre = display_values["materia_prima"]
    tipo = display_values["tipo"].strip().lower()
    proveedor_nombre = display_values["proveedor"]
    trabajador_referencia = display_values["trabajador"]
    motivo = display_values["motivo"].strip() or None
    referencia = display_values["referencia"].strip()

    materia = _find_lookup_value(materia_nombre, materias)
    proveedor = _find_lookup_value(proveedor_nombre, proveedores) if proveedor_nombre else None
    trabajador = _find_lookup_value(trabajador_referencia, trabajadores) if trabajador_referencia else None

    if not materia:
        errors.append(
            _build_catalog_error(
                "materia prima",
                _find_closest_alias(materia_nombre, materia_aliases),
                "La materia prima no existe o no coincide con el catálogo.",
            )
        )

    if proveedor_nombre and not proveedor:
        errors.append(
            _build_catalog_error(
                "proveedor",
                _find_closest_alias(proveedor_nombre, proveedor_aliases),
                "El proveedor indicado no existe o no coincide con el catálogo.",
            )
        )

    if trabajador_referencia and not trabajador:
        errors.append(
            _build_catalog_error(
                "trabajador",
                _find_closest_alias(trabajador_referencia, trabajador_aliases),
                "El trabajador indicado no existe o no coincide con el catálogo.",
            )
        )

    if tipo not in {"entrada", "salida", "ajuste"}:
        errors.append("El tipo debe ser entrada, salida o ajuste.")

    try:
        cantidad = _parse_decimal(raw_row.get("cantidad"))
    except ValueError as exc:
        cantidad = None
        errors.append(str(exc))

    try:
        fecha = _parse_date(raw_row.get("fecha"))
    except ValueError as exc:
        fecha = None
        errors.append(str(exc))

    if not referencia:
        errors.append("La referencia es obligatoria.")

    if cantidad is not None:
        if tipo in {"entrada", "salida"} and cantidad <= 0:
            errors.append("La cantidad debe ser mayor que cero para entradas y salidas.")
        if tipo == "ajuste" and cantidad == 0:
            errors.append("El ajuste no puede ser cero.")

    payload = None
    if not errors:
        payload = {
            "materia_prima_id": materia.id,
            "proveedor_id": proveedor.id if tipo == "entrada" and proveedor else None,
            "trabajador_produccion_id": trabajador.id if tipo == "salida" and trabajador else None,
            "tipo": tipo,
            "cantidad": str(cantidad),
            "fecha": fecha,
            "motivo": motivo,
            "referencia": referencia,
        }
        # El serializer centraliza reglas de negocio y evita duplicar validaciones aquí.
        serializer = MovimientoInventarioSerializer(data=payload)
        if not serializer.is_valid():
            errors.extend(_flatten_serializer_errors(serializer.errors))

    preview_row = {
        "row_number": row_number,
        "status": "error" if errors else "valid",
        "errors": errors,
        "values": display_values,
        "resolved": {
            "materia_prima": materia.nombre if materia else None,
            "proveedor": proveedor.nombre if proveedor else None,
            "trabajador": trabajador.nombre if trabajador else None,
        },
    }
    return preview_row, payload if not errors else None


def analyze_movimiento_import(file_name: str, file_bytes: bytes) -> tuple[dict[str, object], list[dict[str, object]]]:
    """Analiza el archivo completo y produce preview + payloads válidos."""

    # Preview y commit comparten exactamente este análisis para evitar diferencias.
    source_rows, header_observations = _read_source_rows(file_name, file_bytes)
    materias, materia_aliases, proveedores, proveedor_aliases, trabajadores, trabajador_aliases = _build_lookup_maps()
    preview_rows: list[dict[str, object]] = []
    payloads: list[dict[str, object]] = []
    file_errors: list[str] = []

    if not source_rows:
        file_errors.append("El archivo no contiene filas de datos para importar.")

    for row_number, raw_row in source_rows:
        preview_row, payload = _build_row_preview(
            row_number,
            raw_row,
            materias,
            materia_aliases,
            proveedores,
            proveedor_aliases,
            trabajadores,
            trabajador_aliases,
        )
        preview_rows.append(preview_row)
        if payload:
            payloads.append(payload)

    invalid_rows = sum(1 for row in preview_rows if row["status"] == "error")
    preview = {
        "file_name": file_name,
        "file_errors": file_errors,
        "header_observations": header_observations,
        "total_rows": len(preview_rows),
        "valid_rows": len(payloads),
        "invalid_rows": invalid_rows,
        "can_import": not file_errors and invalid_rows == 0 and bool(payloads),
        "rows": preview_rows,
    }
    return preview, payloads


def build_movimiento_import_preview(file_name: str, file_bytes: bytes) -> dict[str, object]:
    """Endpoint helper: solo genera la vista previa, sin persistir cambios."""

    preview, _payloads = analyze_movimiento_import(file_name, file_bytes)
    return preview


@transaction.atomic
def import_movimientos(file_name: str, file_bytes: bytes, user: AppUser | None) -> dict[str, object]:
    """Importa el lote completo de movimientos con rollback ante cualquier error."""

    preview, payloads = analyze_movimiento_import(file_name, file_bytes)

    if not preview["can_import"]:
        return {
            "created_count": 0,
            "preview": preview,
        }

    created_ids: list[int] = []

    # Si una fila falla al guardar, la transacción revierte todo el lote.
    for payload in payloads:
        serializer = MovimientoInventarioSerializer(data=payload)
        serializer.is_valid(raise_exception=True)
        instance = serializer.save(usuario=user)
        created_ids.append(instance.id)

    return {
        "created_count": len(created_ids),
        "created_ids": created_ids,
        "preview": preview,
    }


def build_movimiento_import_template_response() -> HttpResponse:
    """Genera una plantilla Excel de referencia para carga masiva."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "movimientos"
    worksheet.append(TEMPLATE_HEADERS)
    worksheet.append([
        "Resina Poliester",
        "entrada",
        "120.5",
        "2026-05-02",
        "Proveedor Alfa",
        "",
        "Compra inicial de materia prima",
        "ENT-0001",
    ])
    worksheet.append([
        "Resina Poliester",
        "salida",
        "25",
        "2026-05-03",
        "",
        "TR-001",
        "Consumo para orden de producción",
        "SAL-0001",
    ])
    worksheet.append([
        "Catalizador MEKP",
        "ajuste",
        "-2",
        "2026-05-04",
        "",
        "",
        "Corrección por conteo físico",
        "AJU-0001",
    ])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_movimientos_inventario.xlsx"'
    return response